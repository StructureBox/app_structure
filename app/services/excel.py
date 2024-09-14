# app/services/excel.py

import io
import logging
from datetime import datetime
from fastapi import HTTPException

from openpyxl import load_workbook
from services.supabase_utils import (
    ensure_extension,
    generate_supabase_url,
    upload_file_to_supabase,
    generate_download_link,
    download_file_from_url,
)
from models.excel_models import template_cell_map

def get_excel_template(template_name: str) -> io.BytesIO:
    """
    SupabaseからExcelテンプレートを取得する関数。
    """
    logging.debug(f"Requesting signed URL for {template_name} from Supabase.")

    # 拡張子を確実に付ける
    template_file_name = ensure_extension(template_name, ".xlsx")

    # 'excel_templates' バケットからサイン付きURLを生成
    signed_url = generate_supabase_url(template_file_name, bucket_name="excel_templates")

    # サイン付きURLを使ってテンプレートをダウンロード
    return download_file_from_url(signed_url)

def edit_excel_template(template_data: io.BytesIO, template_name: str, data: dict) -> io.BytesIO:
    """
    Excelテンプレートを編集し、データを埋め込む関数。
    """
    wb = load_workbook(template_data)
    ws = wb.active

    # テンプレートに基づいたセルマッピングを取得
    cell_map = template_cell_map.get(template_name)
    if not cell_map:
        raise ValueError(f"Template '{template_name}' not found or not supported")

    # データに基づいてセルに値を入力
    for field, cell_address in cell_map.items():
        if field in data:
            ws[cell_address] = data[field]

    # 編集後のExcelファイルをバイトストリームに保存
    edited_excel = io.BytesIO()
    wb.save(edited_excel)
    edited_excel.seek(0)

    return edited_excel

def process_excel_template(template_name: str, input_data: dict) -> dict:
    """
    Excelテンプレートを処理し、編集後のファイルのダウンロードリンクを生成する関数。
    """
    try:
        # テンプレートを取得
        template = get_excel_template(template_name)

        # テンプレートを編集
        edited_excel = edit_excel_template(template, template_name, input_data)

        # 一意のファイル名を生成
        timestamp_str = datetime.now().strftime("%Y%m%d%H%M%S")
        excel_file_name = f"{template_name}_{timestamp_str}.xlsx"

        # 'edited-zumen' バケットにファイルをアップロード
        upload_file_to_supabase(excel_file_name, edited_excel, bucket_name="edited-zumen")

        # ダウンロードリンクを生成
        excel_download_url = generate_download_link(excel_file_name, bucket_name="edited-zumen")

        return {"excel_download_url": excel_download_url}
    except Exception as e:
        logging.error(f"Error processing Excel template: {str(e)}")
        raise HTTPException(status_code=500, detail="An error occurred while processing the Excel template.")
