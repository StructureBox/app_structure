import os
from openpyxl import load_workbook
import io
import logging
from fastapi import HTTPException


def get_excel_template() -> io.BytesIO:
    """
    ローカルファイルからエクセルテンプレートを取得する関数。
    """
    template_path = "temp/template.xlsx"
    if not os.path.exists(template_path):
        raise HTTPException(status_code=404, detail="Template not found")
    
    with open(template_path, "rb") as f:
        return io.BytesIO(f.read())



def edit_excel_template(template: io.BytesIO, **data) -> io.BytesIO:
    """
    受け取ったデータに基づいてエクセルテンプレートを編集する関数。
    """
    wb = load_workbook(template)
    ws = wb.active

    # データの編集（セルのアドレスは仮定です。実際のテンプレートに合わせて修正してください）
    ws["B47"] = data["architect_number"]
    ws["B48"] = data["architect_name"]
    ws["B49"] = data["office_number"]
    ws["B50"] = data["address"]
    ws["B51"] = data["phone_number"]
    ws["B52"] = data["client_name"]
    ws["B53"] = data["building_location"]
    ws["B54"] = data["building_name"]
    ws["B55"] = data["building_usage"]
    ws["B56"] = data["building_area"]
    ws["B57"] = data["total_area"]
    ws["B58"] = data["max_height"]
    ws["B59"] = data["eaves_height"]
    ws["B60"] = data["above_ground_floors"]
    ws["B61"] = data["underground_floors"]
    ws["B62"] = data["structure_type"]

    # 編集後のエクセルファイルをバイトストリームに保存
    edited_excel = io.BytesIO()
    wb.save(edited_excel)
    edited_excel.seek(0)

    return edited_excel


if __name__ == "__main__":
    # ロギングの設定
    logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

    logging.debug("Starting the script.")
    try:
        # get_excel_templateのテスト実行
        logging.debug("Running get_excel_template to test the function.")
        template_data = get_excel_template()
        logging.info("Template successfully loaded and tested.")
    except HTTPException as e:
        logging.error(f"HTTPException occurred: {e.detail}")
    except Exception as e:
        logging.error(f"An unexpected error occurred: {str(e)}")

    logging.debug("Starting the script.")
    try:
        # get_excel_templateのテスト実行
        logging.debug("Running get_excel_template to test the function.")
        template_data = get_excel_template()
        logging.info("Template successfully loaded and tested.")

        # edit_excel_templateのテスト実行
        logging.debug("Running edit_excel_template to test the function.")
        test_data = {
            "architect_number": "12345",
            "architect_name": "John Doe",
            "office_number": "67890",
            "address": "123 Main St",
            "phone_number": "555-1234",
            "client_name": "Client Name",
            "building_location": "Location",
            "building_name": "Building",
            "building_usage": "Commercial",
            "building_area": 5000.0,
            "total_area": 15000.0,
            "max_height": 50.0,
            "eaves_height": 45.0,
            "above_ground_floors": 5,
            "underground_floors": 1,
            "structure_type": "Steel",
        }
        edited_excel = edit_excel_template(template_data, **test_data)
        logging.info("Excel template successfully edited and tested.")

        # 編集後のエクセルファイルをローカルに保存して確認する（オプション）
        with open("temp/edited_template.xlsx", "wb") as f:
            f.write(edited_excel.getvalue())
        logging.debug("Edited template saved as 'temp/edited_template.xlsx' for verification.")

    except HTTPException as e:
        logging.error(f"HTTPException occurred: {e.detail}")
    except Exception as e:
        logging.error(f"An unexpected error occurred: {str(e)}")
