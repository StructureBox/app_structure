import io

# import os
import logging

from fastapi import HTTPException
from openpyxl import load_workbook
from utils.supabase_utils import generate_supabase_file_url, download_excel_from_url
from models.excel_models import template_cell_map


def get_excel_template(file_name: str) -> io.BytesIO:
    """
    Supabaseからサイン付きURLを使ってExcelテンプレートを取得する関数
    """
    logging.debug(f"Requesting signed URL for {file_name} from Supabase.")

    # 有効期限付きのサイン付きURLを取得（10分間有効）
    signed_url = generate_supabase_file_url(file_name)

    # サイン付きURLを使ってExcelファイルをダウンロード
    return download_excel_from_url(signed_url)


def edit_excel_template(template: io.BytesIO, template_name: str, data: dict) -> io.BytesIO:
    """
    汎用的なExcelテンプレート編集関数。
    テンプレート名に基づいてセルのマッピングを取得し、受け取ったデータを編集。
    """
    wb = load_workbook(template)
    ws = wb.active

    # テンプレートに基づいたセルマッピングを取得
    cell_map = template_cell_map.get(template_name)
    if not cell_map:
        raise ValueError(f"Template '{template_name}' not found or not supported")

    # データに基づいてセルに値を入力
    for field, cell_address in cell_map.items():
        if field in data:
            ws[cell_address] = data[field]

    # 編集後のエクセルファイルをバイトストリームに保存
    edited_excel = io.BytesIO()
    wb.save(edited_excel)
    edited_excel.seek(0)

    return edited_excel


def edit_excel_safety_certificate(template: io.BytesIO, **data) -> io.BytesIO:
    """
    受け取ったデータに基づいてエクセルテンプレートを編集する関数。
    """
    wb = load_workbook(template)
    ws = wb.active

    # データの編集（セルのアドレスは仮定です。実際のテンプレートに合わせて修正してください）
    ws["B47"] = data["architect_number"]
    ws["B48"] = data["architect_name"]
    ws["B49"] = data["office_number"]
    ws["B50"] = data["address"]
    ws["B51"] = data["phone_number"]
    ws["B52"] = data["client_name"]
    ws["B53"] = data["building_location"]
    ws["B54"] = data["building_name"]
    ws["B55"] = data["building_usage"]
    ws["B56"] = data["building_area"]
    ws["B57"] = data["total_area"]
    ws["B58"] = data["max_height"]
    ws["B59"] = data["eaves_height"]
    ws["B60"] = data["above_ground_floors"]
    ws["B61"] = data["underground_floors"]
    ws["B62"] = data["structure_type"]

    # 編集後のエクセルファイルをバイトストリームに保存
    edited_excel = io.BytesIO()
    wb.save(edited_excel)
    edited_excel.seek(0)

    return edited_excel


if __name__ == "__main__":
    # ロギングの設定
    logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")

    logging.debug("Starting the script.")
    try:
        # get_excel_templateのテスト実行
        logging.debug("Running get_excel_template to test the function.")
        template_data = get_excel_template()
        logging.info("Template successfully loaded and tested.")
    except HTTPException as e:
        logging.error(f"HTTPException occurred: {e.detail}")
    except Exception as e:
        logging.error(f"An unexpected error occurred: {str(e)}")

    logging.debug("Starting the script.")
    try:
        # get_excel_templateのテスト実行
        logging.debug("Running get_excel_template to test the function.")
        template_data = get_excel_template()
        logging.info("Template successfully loaded and tested.")

        # edit_excel_templateのテスト実行
        logging.debug("Running edit_excel_template to test the function.")
        test_data = {
            "architect_number": "12345",
            "architect_name": "John Doe",
            "office_number": "67890",
            "address": "123 Main St",
            "phone_number": "555-1234",
            "client_name": "Client Name",
            "building_location": "Location",
            "building_name": "Building",
            "building_usage": "Commercial",
            "building_area": 5000.0,
            "total_area": 15000.0,
            "max_height": 50.0,
            "eaves_height": 45.0,
            "above_ground_floors": 5,
            "underground_floors": 1,
            "structure_type": "Steel",
        }
        edited_excel = edit_excel_template(template_data, **test_data)
        logging.info("Excel template successfully edited and tested.")

        # 編集後のエクセルファイルをローカルに保存して確認する（オプション）
        with open("temp/edited_template.xlsx", "wb") as f:
            f.write(edited_excel.getvalue())
        logging.debug("Edited template saved as 'temp/edited_template.xlsx' for verification.")

    except HTTPException as e:
        logging.error(f"HTTPException occurred: {e.detail}")
    except Exception as e:
        logging.error(f"An unexpected error occurred: {str(e)}")
